# Zakupi.py
# Streamlit app: obrada ugovora o zakupu (SKENIRANI PDF -> OCR -> Zakup.xlsx)
#
# Pokretanje (BAT) - OBAVEZNO:
#   C:\Python\Python311\python.exe -m streamlit run C:\Automatika\Python\Streamlit\Zakupi.py
#
# Paketi:
#   pip install streamlit pdfplumber openpyxl pytesseract pdf2image pillow

from __future__ import annotations

import re
import time
import tempfile
from pathlib import Path
from typing import Dict, Any, Optional, List, Tuple

import streamlit as st
import pdfplumber
from openpyxl import Workbook, load_workbook

# OCR
OCR_AVAILABLE = True
try:
    import pytesseract
    from pdf2image import convert_from_path
except Exception:
    OCR_AVAILABLE = False


# =========================================================
# KONFIG (default putanje)
# =========================================================
#DEFAULT_PDF_DIR = r"C:\Users\hr.mdrauto\OneDrive - Inter Cars S.A\Documents\ZAKUPI"
#DEFAULT_EXCEL_PATH = r"C:\Users\hr.mdrauto\OneDrive - Inter Cars S.A\Documents\Zakupi_obrada\Zakup.xlsx"

DEFAULT_PDF_DIR = r"C:\Users\hr.mdrauto\OneDrive - Inter Cars S.A\Documents\Zakupi_ugovori"
DEFAULT_EXCEL_PATH = r"C:\Users\hr.mdrauto\OneDrive - Inter Cars S.A\Documents\Zakupi_obrada\Zakup.xlsx"

DEFAULT_SHEET = "Zakupi"

TESSERACT_CMD = r"C:\Program Files\Tesseract-OCR\tesseract.exe"
TESSERACT_LANG = "hrv"          # ili "hrv+eng"
POPPLER_BIN = r"C:\Python\poppler\Library\bin"  # promijeni ako treba

# svi dokumenti su skenirani -> OCR je primarni
OCR_DPI = 300
OCR_PSM_CONFIGS = ["--psm 6", "--psm 11"]  # 6=blok teksta, 11=sparse (naslovi)

# za robustnost naslova/ključnih dijelova
HEAD_CHARS = 2500


# =========================================================
# LOCATION_MAP (Hxx -> podaci)
# NAMJERNO PRESKOČENO: ubaci svoj mapping ovdje
# =========================================================
LOCATION_MAP: Dict[str, Dict[str, str]] = {
    "H00": {"PARTNER":"INTER CARS d.o.o.","Adresa partnera":"Zelena Aleja 45, Vukovina","Adresa poslovnice":"Zelena Aleja 45, Vukovina","Poslovnica":"LOGISTIČKI CENTAR","Vlasnik":"Mirko Rugle","VODITELJ":"Mirko Rugle","RVP":""},
    "H02": {"PARTNER":"M.J. CARS d.o.o.","Adresa partnera":"Južna ulica VIII odvojak br. 4, 10000 Zagreb","Adresa poslovnice":"Kovinska 9b, 10090 Zagreb","Poslovnica":"ZAGREB JANKOMIR","Vlasnik":"Mario Jurišić","VODITELJ":"Mario Jurišić","RVP":"Ivan Zidar"},
    "H03": {"PARTNER":"KNEŽEVIĆ CAR PARTS d.o.o","Adresa partnera":"Zelinska 27, 10360 Sesvete","Adresa poslovnice":"Savska cesta 114, 10360 Sesvete","Poslovnica":"ZAGREB SESVETE","Vlasnik":"Goran Knežević","VODITELJ":"Goran Knežević","RVP":"Ivan Zidar"},
    "H04": {"PARTNER":"KNEŽEVIĆ CAR PARTS d.o.o","Adresa partnera":"Zelinska 27, 10360 Sesvete","Adresa poslovnice":"Savska cesta 114, 10360 Sesvete","Poslovnica":"ZAGREB VELIKA GORICA","Vlasnik":"Goran Knežević","VODITELJ":"Goran Knežević","RVP":"Ivan Zidar"},
    "H06": {"PARTNER":"M.J. CARS d.o.o.","Adresa partnera":"Južna ulica VIII odvojak br. 4, 10000 Zagreb","Adresa poslovnice":"Kovinska 9b, 10090 Zagreb","Poslovnica":"ZAGREB DUBRAVA","Vlasnik":"Mario Jurišić","VODITELJ":"Goran Nemec","RVP":"Ivan Zidar"},
    "H11": {"PARTNER":"POSILOVIĆ D.O.O.","Adresa partnera":"Josipa Juraja Strossmayera 118, 44000 Sisak","Adresa poslovnice":"Zagrebačka ulica 46, 44000 Sisak","Poslovnica":"SISAK","Vlasnik":"Stjepan Posilović","VODITELJ":"Stjepan Posilović","RVP":"Ivan Zidar"},
    "H12": {"PARTNER":"M.J. CARS d.o.o.","Adresa partnera":"Južna ulica VIII odvojak br. 4, 10000 Zagreb","Adresa poslovnice":"Kovinska 9b, 10090 Zagreb","Poslovnica":"ZADAR","Vlasnik":"Mario Jurišić","VODITELJ":"Mario Jurišić","RVP":"Mario Milković"},
    "H13": {"PARTNER":"HERBI d.o.o.","Adresa partnera":"Marčelji 54, 51216 Viškovo","Adresa poslovnice":"Marčelji 54, 51000 Rijeka","Poslovnica":"RIJEKA","Vlasnik":"Eleonora Marohnić","VODITELJ":"Eleonora Marohnić","RVP":"Davor Fogec"},
    "H14": {"PARTNER":"KA PROM d.o.o.","Adresa partnera":"Mlinska 1c, 43500 Daruvar","Adresa poslovnice":"Ulica branitelja Orion 2, 42000 Varaždin","Poslovnica":"VARAŽDIN","Vlasnik":"Marijo Kanjuh","VODITELJ":"Marijo Kanjuh","RVP":"Mario Milković"},
    "H15": {"PARTNER":"AUTO LOGISTIKA D.O.O.","Adresa partnera":"Močile 15, 48000 Koprivnica","Adresa poslovnice":"Ulica Josipa Bukovčana 14, 48000 Koprivnica","Poslovnica":"KOPRIVNICA","Vlasnik":"Karlo Babić","VODITELJ":"Karlo Babić","RVP":"Davor Fogec"},
    "H16": {"PARTNER":"KA PROM d.o.o.","Adresa partnera":"Mlinska 1c, 43500 Daruvar","Adresa poslovnice":"Ulica branitelja Orion 2, 42000 Varaždin","Poslovnica":"ČAKOVEC","Vlasnik":"Marijo Kanjuh","VODITELJ":"Marijo Kanjuh","RVP":"Mario Milković"},
    "H17": {"PARTNER":"AUTOPULS vlsnik Nenad Bilokapić","Adresa partnera":"Slavonski Brod, Rudine 17","Adresa poslovnice":"Dubrovačka 65a, 34000 Požega","Poslovnica":"SLAVONSKI BROD","Vlasnik":"Davor Bilokapić","VODITELJ":"Davor Bilokapić","RVP":"Zoran Prolić"},
    "H18": {"PARTNER":"T.L.P. DISTRIBUCIJA d.o.o.","Adresa partnera":"Sime Lončara 38, Mirkovci, 321000 Vinkovci","Adresa poslovnice":"Bana Josipa Jelačića 107B, 32100 Vinkovci","Poslovnica":"VINKOVCI","Vlasnik":"Ivor Ređep","VODITELJ":"Ivor Ređep","RVP":"Davor Fogec"},
    "H19": {"PARTNER":"ASTRALIS-MEDIA d.o.o.","Adresa partnera":"Ulica Trzunove pećine 9, 10000 Zagreb","Adresa poslovnice":"Zagrebačka 15/I, 47000 Karlovac","Poslovnica":"KARLOVAC","Vlasnik":"Bruno Osredečki","VODITELJ":"Bruno Osredečki","RVP":"Mario Milković"},
    "H20": {"PARTNER":"AUTO LOGISTIKA D.O.O.","Adresa partnera":"Močile 15, 48000 Koprivnica","Adresa poslovnice":"Ulica Josipa Bukovčana 14, 48000 Koprivnica","Poslovnica":"VIROVITICA","Vlasnik":"Karlo Babić","VODITELJ":"Karlo Babić","RVP":"Davor Fogec"},
    "H21": {"PARTNER":"NACIONAL LOGISTICS d.o.o.","Adresa partnera":"Bartula Kašića 32, 31000 Osijek","Adresa poslovnice":"Kudeljarska 13, 32000 Vukovar","Poslovnica":"VUKOVAR","Vlasnik":"Davorin Ljubojević","VODITELJ":"Denis Ljubojević","RVP":"Zoran Prolić"},
    "H22": {"PARTNER":"POSILOVIĆ D.O.O.","Adresa partnera":"Josipa Juraja Strossmayera 118, 44000 Sisak","Adresa poslovnice":"Zagrebačka ulica 46, 44000 Sisak","Poslovnica":"KUTINA","Vlasnik":"Stjepan Posilović","VODITELJ":"Stjepan Posilović","RVP":"Ivan Zidar"},
    "H23": {"PARTNER":"KA PROM d.o.o.","Adresa partnera":"Mlinska 1c, 43500 Daruvar","Adresa poslovnice":"Ulica branitelja Orion 2, 42000 Varaždin","Poslovnica":"BJELOVAR","Vlasnik":"Marijo Kanjuh","VODITELJ":"Marijo Kanjuh","RVP":"Mario Milković"},
    "H24": {"PARTNER":"AUTOPULS vlsnik Nenad Bilokapić","Adresa partnera":"Slavonski Brod, Rudine 17","Adresa poslovnice":"Dubrovačka 65a, 34000 Požega","Poslovnica":"POŽEGA","Vlasnik":"Davor Bilokapić","VODITELJ":"Darko Lucić","RVP":"Zoran Prolić"},
    "H25": {"PARTNER":"NACIONAL LOGISTICS d.o.o.","Adresa partnera":"Bartula Kašića 32, 31000 Osijek","Adresa poslovnice":"Kudeljarska 13, 32000 Vukovar","Poslovnica":"OSIJEK","Vlasnik":"Davorin Ljubojević","VODITELJ":"Davorin Ljubojević","RVP":"Zoran Prolić"},
    "H26": {"PARTNER":"E.C. PREMIUM AUTO d.o.o.","Adresa partnera":"Buonarrotijeva ulica 1, 52100 Pula","Adresa poslovnice":"Partizanski put 132, 52100 Pula","Poslovnica":"PULA","Vlasnik":"Edvard Cerkovnik","VODITELJ":"Edvard Cerkovnik","RVP":"Davor Fogec"},
    "H27": {"PARTNER":"M.J. CARS d.o.o.","Adresa partnera":"Južna ulica VIII odvojak br. 4, 10000 Zagreb","Adresa poslovnice":"Kovinska 9b, 10090 Zagreb","Poslovnica":"ŠIBENIK","Vlasnik":"Mario Jurišić","VODITELJ":"Ante Munjiza","RVP":"Mario Milković"},
    "H28": {"PARTNER":"E.C. PREMIUM AUTO d.o.o.","Adresa partnera":"Buonarrotijeva ulica 1, 52100 Pula","Adresa poslovnice":"Partizanski put 132, 52100 Pula","Poslovnica":"POREČ","Vlasnik":"Edvard Cerkovnik","VODITELJ":"Matko Lukić","RVP":"Davor Fogec"},
    "H29": {"PARTNER":"AUTOPULS vlsnik Nenad Bilokapić","Adresa partnera":"Slavonski Brod, Rudine 17","Adresa poslovnice":"Dubrovačka 65a, 34000 Požega","Poslovnica":"NOVA GRADIŠKA","Vlasnik":"Davor Bilokapić","VODITELJ":"Marijo Pereković","RVP":"Zoran Prolić"},
    "H30": {"PARTNER":"CAR CRAFT d.o.o.","Adresa partnera":"Pazdigradska 13, 21000 Split","Adresa poslovnice":"IV.Gardijske brigade 30, 21000 Split","Poslovnica":"SPLIT","Vlasnik":"Hrvoje Štetić","VODITELJ":"Hrvoje Štetić","RVP":"Zoran Prolić"},
    "H31": {"PARTNER":"HERBI d.o.o.","Adresa partnera":"Marčelji 54, 51216 Viškovo","Adresa poslovnice":"Marčelji 54, 51000 Rijeka","Poslovnica":"CRIKVENICA","Vlasnik":"Eleonora Marohnić","VODITELJ":"Eleonora Marohnić","RVP":"Davor Fogec"},
    "H32": {"PARTNER":"ASTRALIS-MEDIA d.o.o.","Adresa partnera":"Ulica Trzunove pećine 9, 10000 Zagreb","Adresa poslovnice":"Zagrebačka 15/I, 47000 Karlovac","Poslovnica":"ZABOK","Vlasnik":"Bruno Osredečki","VODITELJ":"Bruno Osredečki","RVP":"Mario Milković"},
    "H33": {"PARTNER":"CAR CRAFT d.o.o.","Adresa partnera":"Pazdigradska 13, 21000 Split","Adresa poslovnice":"IV.Gardijske brigade 30, 21000 Split","Poslovnica":"KAŠTELA","Vlasnik":"Hrvoje Štetić","VODITELJ":"Stipe Jureta","RVP":"Zoran Prolić"},
    "H34": {"PARTNER":"NACIONAL LOGISTICS d.o.o.","Adresa partnera":"Bartula Kašića 32, 31000 Osijek","Adresa poslovnice":"Kudeljarska 13, 32000 Vukovar","Poslovnica":"ĐAKOVO","Vlasnik":"Davorin Ljubojević","VODITELJ":"Davorin Ljubojević","RVP":"Zoran Prolić"},
}

from urllib.parse import quote

#ONEDRIVE_LOCAL_ROOT = Path(
#    r"C:\Users\hr.mdrauto\OneDrive - Inter Cars S.A\Documents"
#)

ONEDRIVE_LOCAL_ROOT = Path.home() / "OneDrive - Inter Cars S.A" / "Documents"

ONEDRIVE_WEB_BASE = (
    "https://icars-my.sharepoint.com/"
    "personal/hr_mdrauto_intercars_eu/"
    "Documents/Documents"
)


def onedrive_web_url_for_path(file_path: Path) -> str:
    p = file_path.resolve()

    try:
        rel = p.relative_to(ONEDRIVE_LOCAL_ROOT).as_posix()
    except ValueError:
        return ""

    rel_enc = quote(rel, safe="/")
    return f"{ONEDRIVE_WEB_BASE}/{rel_enc}?web=1"

# =========================================================
# EXCEL HEADERI
# =========================================================
EXCEL_HEADERS: List[str] = [
    "Broj/naziv ugovora","Ugovor / aneks","Partner (podzakupnik)","Ime","Adresa partnera","OIB partnera",
    "odjel kod","Poslovnica","Adresa poslovnice","Naziv zakupodavca","OIB zakupodavca","Adresa sjedišta zakupodavca",
    "z.k.č.br.","z.k.ul","općinski sud","katastarska općina",
    "Datum potpisivanja ugovora /aneksa","Datum stupanja ugovora na snagu","Datum isteka ugovora",
    "Trajanje ugovora u mjesecima","otkaz","datum za slanje maila za pregovaranje",
    "Predmet zakupa","Kvadratura ukupnog prostora u zakupu u m2","Kvadratura skladišta","Kvadratura uredski prostori",
    "Ostale prostorije","Broj parkirnih mjesta","Jedinična cijena/m2 (Bez PDV-a)","AKTUALNA CIJENA ZAKUPA",
    "Ukupna vrijednost ugovora (mjesečni iznos zakupnine)","Ukupna vrijednost ugovora ugovoreno u kunama",
    "Indeksacija","iznos ineksacija godišnje","Režijski troškovi uključeni?","Korisnik ugovora",
    "STATUS UGOVORA/ANEKSA","DATUM OTKAZIVANJA / ISTEKA","Podzakup","Datum isteka","Suglasnost za podzakup",
    "Zadužnice","RVP","APPOINTMENT - OUTLOOK KALENDAR","ENERGETSKI CERTIFIKAT",
]

# =========================================================
# POMOĆNE FUNKCIJE (Excel)
# =========================================================
def excel_locked_message(path: Path) -> None:
    st.error(
        f"❌ Excel datoteka je trenutno otvorena: **{path}**\n\n"
        f"Zatvori datoteku u Excelu i pokušaj ponovno."
    )
    st.stop()

def safe_load_workbook(path: Path, attempts: int = 3, delay: int = 2):
    for _ in range(attempts):
        try:
            return load_workbook(path)
        except PermissionError:
            time.sleep(delay)
    excel_locked_message(path)

def ensure_excel(excel_path: Path, sheet_name: str) -> None:
    excel_path.parent.mkdir(parents=True, exist_ok=True)

    if not excel_path.exists():
        wb = Workbook()
        try:
            ws = wb.active
            ws.title = sheet_name
            ws.append(EXCEL_HEADERS)
            wb.save(excel_path)
        except PermissionError:
            excel_locked_message(excel_path)
        finally:
            wb.close()
        return

    wb = safe_load_workbook(excel_path)
    try:
        if sheet_name not in wb.sheetnames:
            ws = wb.create_sheet(sheet_name)
            ws.append(EXCEL_HEADERS)
        else:
            ws = wb[sheet_name]
            existing = [ws.cell(1, c).value for c in range(1, len(EXCEL_HEADERS) + 1)]
            if existing != EXCEL_HEADERS:
                if ws.max_row < 1:
                    ws.append(EXCEL_HEADERS)
                else:
                    ws.insert_rows(1)
                    for i, h in enumerate(EXCEL_HEADERS, start=1):
                        ws.cell(1, i).value = h

        try:
            wb.save(excel_path)
        except PermissionError:
            excel_locked_message(excel_path)
    finally:
        wb.close()

def _col_index(header: str) -> int:
    return EXCEL_HEADERS.index(header) + 1  # 1-based

def append_row(excel_path: Path, sheet_name: str, row_dict: Dict[str, Any], hyperlinks: Optional[Dict[str, str]] = None) -> int:
    ensure_excel(excel_path, sheet_name)

    wb = safe_load_workbook(excel_path)
    try:
        ws = wb[sheet_name]
        ws.append([row_dict.get(h, "") for h in EXCEL_HEADERS])
        idx = ws.max_row

        if hyperlinks:
            for header, url in hyperlinks.items():
                if not url or header not in EXCEL_HEADERS:
                    continue
                c = ws.cell(row=idx, column=_col_index(header))
                if c.value in (None, ""):
                    c.value = row_dict.get(header, "")
                c.hyperlink = url
                c.style = "Hyperlink"

        wb.save(excel_path)
        return idx
    except PermissionError:
        excel_locked_message(excel_path)
    finally:
        wb.close()

# =========================================================
# POMOĆNE FUNKCIJE (PDF / OCR)
# =========================================================
def normalize_spaces(s: str) -> str:
    s = s.replace("\u00a0", " ")
    s = re.sub(r"[ \t]+", " ", s)
    return s.strip()

def pdf_text_extract(pdf_path: Path) -> str:
    texts = []
    with pdfplumber.open(str(pdf_path)) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            texts.append(t)
    return normalize_spaces("\n".join(texts))

def read_pdf_scanned(pdf_path: Path) -> str:
    parts: List[str] = []
    try:
        t = pdf_text_extract(pdf_path)
        if t:
            parts.append(t)
    except Exception:
        pass

    if not OCR_AVAILABLE:
        return normalize_spaces("\n".join(parts))

    if TESSERACT_CMD:
        pytesseract.pytesseract.tesseract_cmd = TESSERACT_CMD

    try:
        images = convert_from_path(str(pdf_path), dpi=OCR_DPI, poppler_path=POPPLER_BIN)
    except Exception:
        images = []

    ocr_texts: List[str] = []
    for img in images:
        for cfg in OCR_PSM_CONFIGS:
            try:
                ocr_texts.append(pytesseract.image_to_string(img, lang=TESSERACT_LANG, config=cfg))
            except Exception:
                pass

    parts.append("\n".join(ocr_texts))
    return normalize_spaces("\n".join(parts))

# =========================================================
# FORMAT HELPERS (adresa)
# =========================================================
def clean_brackets(s: str) -> str:
    if not s:
        return ""
    s = re.sub(r"\(\s*\)", "", s)
    s = s.replace("(", " ").replace(")", " ")
    s = re.sub(r"\s+", " ", s).strip(" ,;:-–—")
    return s


def normalize_address_case(s: str) -> str:
    """
    Normalizira velika/mala slova:
    - Title Case po tokenima
    - čuva brojeve, poštanske brojeve, rimske brojeve (IV.)
    - čuva kratice od 2-5 slova ako su full uppercase
    - ALI: riječi poput ULICA/CESTA/TRG/PUT/NASELJE nisu kratice -> pretvori u Title Case
    """
    if not s:
        return ""

    s = re.sub(r"\s+", " ", s.strip())
    tokens = s.split(" ")

    keep_upper = {"BB", "ZK", "OIB", "DD"}
    not_acronyms = {"ULICA", "CESTA", "TRG", "PUT", "NASELJE", "BROJ", "BR"}

    out: List[str] = []
    for raw in tokens:
        if not raw:
            out.append(raw)
            continue

        raw_stripped = raw.strip()

        if re.fullmatch(r"[0-9]+([\/\\\-][0-9]+)?", raw_stripped):
            out.append(raw_stripped)
            continue

        if re.fullmatch(r"\d{5}", raw_stripped):
            out.append(raw_stripped)
            continue

        if re.fullmatch(r"[IVXLCDM]+\.", raw_stripped.upper()):
            out.append(raw_stripped.upper())
            continue

        if raw_stripped.upper().strip(".,") in not_acronyms:
            out.append(raw_stripped[:1].upper() + raw_stripped[1:].lower())
            continue

        if raw_stripped.upper().strip(".,") in keep_upper:
            out.append(raw_stripped.upper())
            continue

        letters_only = re.sub(r"[^A-Za-zČĆŠĐŽčćšđž]", "", raw_stripped)
        if raw_stripped.isupper() and 2 <= len(letters_only) <= 5:
            out.append(raw_stripped)
            continue

        if re.search(r"\d", raw_stripped) and "/" in raw_stripped:
            out.append(raw_stripped)
            continue

        out.append(raw_stripped[:1].upper() + raw_stripped[1:].lower())

    s2 = " ".join(out)
    s2 = re.sub(r"\s*,\s*", ", ", s2)
    s2 = re.sub(r"\s*;\s*", "; ", s2)
    s2 = re.sub(r"\s+", " ", s2).strip(" ,;:-–—")
    return s2


def finalize_address(addr: str) -> str:
    addr = clean_brackets(addr)
    addr = normalize_address_case(addr)
    return addr


# =========================================================
# PARSERI (naslov, zakupodavac, zk, kvadrature, datumi)
# =========================================================
def extract_location_from_filename(filename: str) -> Optional[str]:
    m = re.search(r"\bH\d{2}\b", filename.upper())
    return m.group(0) if m else None


def detect_contract_title(text: str) -> str:
    if not text:
        return ""
    norm = re.sub(r"\s+", " ", text.upper()).replace("0", "O")

    aneks_re = re.compile(
        r"A\W*N\W*E\W*K\W*S\W+U\W*G\W*O\W*V\W*O\W*R\W*(?:A|U)\W+O\W+Z\W*A\W*K\W*U\W*P\W*U",
        re.IGNORECASE
    )
    ugovor_re = re.compile(
        r"U\W*G\W*O\W*V\W*O\W*R\W+O\W+Z\W*A\W*K\W*U\W*P\W*U",
        re.IGNORECASE
    )

    if aneks_re.search(norm):
        return "Aneks ugovora o zakupu"
    if ugovor_re.search(norm):
        return "Ugovor o zakupu"
    return ""


def detect_contract_type(text: str) -> str:
    head = (text or "")[:1200].strip()
    head_up = re.sub(r"\s+", " ", head.upper()).replace("0", "O")
    if re.search(r"\bA\W*N\W*E\W*K\W*S\b", head_up):
        return "Aneks"
    return "Ugovor"


SKLOPLJEN_RE = re.compile(r"\bsklopljen\s+između\s*:\s*", re.IGNORECASE)
ZAKUPO_WORD_RE = re.compile(r"z\W*a\W*k\W*u\W*p\W*o\W*d\W*a\W*v\W*a\W*c", re.IGNORECASE)


def has_zakupodavac_token(s: str) -> bool:
    return bool(ZAKUPO_WORD_RE.search(s or ""))


ROLE_TRIGGER_ZAKUPO = (
    r"(dalje\s+u\s+tekstu\s*:?\s*['\"“”]?\s*zakupodavac\s*['\"“”]?|"
    r"u\s+daljnjem\s+tekstu\s*:?\s*zakupodavac)"
)

OIB_LABEL = r"O\W*[I1l]{1,2}\W*[B8]"
OIB_RE = re.compile(rf"\b{OIB_LABEL}\b\s*[:\-]?\s*([0-9]{{11}})\b", re.IGNORECASE)

LEGAL_SUFFIX_RE = re.compile(
    r"\b("
    r"d\W*o\W*o\W*\.?|"
    r"j\W*d\W*o\W*o\W*\.?|"
    r"d\W*d\W*\.?|"
    r"k\W*d\W*\.?|"
    r"obrt|"
    r"zadruga|"
    r"u\W*d\W*r\W*u\W*g\W*a"
    r")\b",
    re.IGNORECASE
)


def _clean_line(s: str) -> str:
    s = s.replace("\u00a0", " ")
    s = re.sub(r"[ \t]+", " ", s)
    return s.strip(" -–—\t\r\n")


def _strip_role_tail(s: str) -> str:
    s = re.split(r"(dalje\s+u\s+tekstu|u\s+daljnjem\s+tekstu)", s, flags=re.IGNORECASE)[0]
    return s.strip(" ,;:-–—")


def _strip_sklopljen_prefix(s: str) -> str:
    s = SKLOPLJEN_RE.sub("", s).strip()
    return s.strip(" ,;:-–—")


def _name_until_stop(s: str) -> str:
    s = _strip_sklopljen_prefix(s)
    s = _strip_role_tail(s)
    s = re.split(rf"\b{OIB_LABEL}\b|\bzastupan\b|\bzastupano\b|\bkao\b", s, flags=re.IGNORECASE)[0]
    if "," in s:
        s = s.split(",")[0]
    return s.strip(" ,;:-–—")


def cut_company_name_if_applicable(name: str, context: str) -> str:
    if not name:
        return name
    if not has_zakupodavac_token(context):
        return name.strip()

    m = LEGAL_SUFFIX_RE.search(name)
    if not m:
        return name.strip()

    cut = name[:m.end()].strip()
    cut = re.sub(r"\bd\W*o\W*o\W*\.?\b", "d.o.o.", cut, flags=re.IGNORECASE)
    cut = re.sub(r"\bj\W*d\W*o\W*o\W*\.?\b", "j.d.o.o.", cut, flags=re.IGNORECASE)
    cut = re.sub(r"\bd\W*d\W*\.?\b", "d.d.", cut, flags=re.IGNORECASE)
    cut = re.sub(r"\bk\W*d\W*\.?\b", "k.d.", cut, flags=re.IGNORECASE)
    return cut


def _looks_like_address(text: str) -> bool:
    if not text:
        return False
    t = text.strip()

    if re.search(r"\bzastupan\b|\bzastupano\b", t, re.IGNORECASE):
        return False
    if re.search(OIB_LABEL, t, re.IGNORECASE):
        return False
    if re.search(r"\b\d{11}\b", t):
        return False

    has_street_word = bool(re.search(
        r"\bulica\b|\bcesta\b|\btrg\b|\bput\b|\bavenija\b|\bkolodvorska\b|\bnaselje\b",
        t,
        re.IGNORECASE
    ))
    has_zip = bool(re.search(r"\b\d{5}\b", t))
    has_house_no = bool(re.search(r"\b\d{1,4}\b", t)) and (has_street_word or "," in t)

    return has_zip or has_street_word or has_house_no


def _address_after_oib(block: str) -> str:
    b = re.sub(r"\s+", " ", block.strip())
    m = OIB_RE.search(b)
    if not m:
        return ""

    after = b[m.end():].strip(" ,;:-–—")
    after = re.split(
        r"\bzastupan\b|\bzastupano\b|\bkao\b|dalje\s+u\s+tekstu|u\s+daljnjem\s+tekstu",
        after,
        flags=re.IGNORECASE
    )[0].strip(" ,;:-–—")

    if _looks_like_address(after):
        return finalize_address(after)

    before = b[:m.start()].strip(" ,;:-–—")
    parts = [p.strip() for p in before.split(",") if p.strip()]

    if len(parts) >= 2:
        candidate = ", ".join(parts[1:]).strip(" ,;:-–—")
        candidate = re.split(r"\bzastupan\b|\bzastupano\b", candidate, flags=re.IGNORECASE)[0].strip(" ,;:-–—")
        if _looks_like_address(candidate):
            return finalize_address(candidate)

    return ""


def extract_zakupodavac(text: str) -> dict:
    lines = [_clean_line(x) for x in (text or "").splitlines() if _clean_line(x)]

    for i, line in enumerate(lines):
        if SKLOPLJEN_RE.search(line):
            block = line
            if i + 1 < len(lines) and len(_strip_sklopljen_prefix(line)) <= 25:
                block = f"{line} {lines[i+1]}"

            name = _name_until_stop(block)
            name = cut_company_name_if_applicable(name, block)

            mo = OIB_RE.search(block)
            oib = mo.group(1) if mo else ""

            addr = _address_after_oib(block)

            if name:
                return {"name": name, "oib": oib, "address": addr}

    for i in range(len(lines)):
        window = " ".join(lines[i:i + 3])
        if re.search(ROLE_TRIGGER_ZAKUPO, window, flags=re.IGNORECASE):
            prefix = re.split(ROLE_TRIGGER_ZAKUPO, window, flags=re.IGNORECASE)[0]
            prefix = _strip_sklopljen_prefix(prefix)

            mo = OIB_RE.search(window)
            oib = mo.group(1) if mo else ""

            name = _name_until_stop(prefix)
            name = cut_company_name_if_applicable(name, window)

            addr = _address_after_oib(window) or _address_after_oib(prefix)

            if name:
                return {"name": name, "oib": oib, "address": addr}

    for i in range(len(lines)):
        line = lines[i]
        if re.search(r"\bkao\s+zakupodavac\b", line, re.IGNORECASE) or has_zakupodavac_token(line):
            name = _name_until_stop(line)
            name = cut_company_name_if_applicable(name, line)

            mo = OIB_RE.search(line)
            oib = mo.group(1) if mo else ""

            addr = _address_after_oib(line)

            if (not oib or not addr) and i + 1 < len(lines):
                nxt = lines[i + 1]
                if not oib:
                    mo2 = OIB_RE.search(nxt)
                    oib = mo2.group(1) if mo2 else oib
                if not addr:
                    addr = _address_after_oib(nxt) or addr

            if name:
                return {"name": name, "oib": oib, "address": addr}

    return {"name": "", "oib": "", "address": ""}


def extract_zk_ul(text: str) -> str:
    patterns = [
        r"\bz\.?\s*k\.?\s*uložak\s*br\.?\s*[:\-]?\s*([0-9]+)",
        r"\bBroj\s+ZK\s+uloška\s*[:\-]?\s*([0-9]+)",
        r"\bZK\s+uložak\s*[:\-]?\s*([0-9]+)",
        r"\bzk\W*uložak\W*([0-9]+)",
    ]
    for p in patterns:
        m = re.search(p, text, flags=re.IGNORECASE)
        if m:
            return m.group(1).strip()
    return ""


def extract_zk_c_br(text: str) -> str:
    found: List[str] = []

    def _norm_parcel(v: str) -> str:
        v = v.replace(" ", "")
        v = v.replace("\\", "/").replace("-", "/")
        v = v.replace("l", "1").replace("I", "1")
        return v

    patterns_single = [
        r"\bk\W*[čc]\W*br\W*[:\.\-]?\W*([0-9]{1,6}\s*[/\\\-]\s*[0-9]{1,6})",
        r"\bkatastarsk\w*\s+čestic\w*\s+br\W*[:\.\-]?\W*([0-9]{1,6}\s*[/\\\-]\s*[0-9]{1,6})",
    ]
    for p in patterns_single:
        for m in re.finditer(p, text, flags=re.IGNORECASE):
            found.append(_norm_parcel(m.group(1)))

    for m in re.finditer(r"\bčest\W*zem\W*\.?\b(.{0,140})", text, flags=re.IGNORECASE | re.DOTALL):
        window = m.group(1)
        for v in re.findall(r"([0-9]{1,6}\s*[/\\\-]\s*[0-9]{1,6})", window):
            found.append(_norm_parcel(v))

    uniq: List[str] = []
    seen = set()
    for v in found:
        if v and v not in seen:
            seen.add(v)
            uniq.append(v)

    return "; ".join(uniq)


def extract_opcinski_sud(text: str) -> str:
    patterns = [
        r"\bOpćinski\s+građanski\s+sud\s+u\s+([A-Za-zČĆŠĐŽčćšđž]+)",
        r"\bOpćinski\s+sud\s+u\s+([A-Za-zČĆŠĐŽčćšđž]+)",
        r"\bOpćinskog\s+suda\s+u\s+([A-Za-zČĆŠĐŽčćšđž]+)",
        r"\bNadležnog\s+suda\s+u\s+([A-Za-zČĆŠĐŽčćšđž]+)",
    ]
    for p in patterns:
        m = re.search(p, text, flags=re.IGNORECASE)
        if m:
            city = re.sub(r"[^A-Za-zČĆŠĐŽčćšđž]", "", m.group(1).strip())
            city = city[:1].upper() + city[1:].lower() if city else city
            return f"Općinski sud u {city}"
    return ""


def extract_parking_first_item(text: str) -> str:
    m = re.search(r"\buz\s+([0-9]+)\s+parkirna\s+mjesta\b", text, flags=re.IGNORECASE)
    if m:
        return m.group(1).strip()
    m = re.search(r"\b([0-9]+)\s+parkirnih\s+mjesta\b", text, flags=re.IGNORECASE)
    return m.group(1).strip() if m else ""


AREA_UNIT_RE = r"(?:m\s*(?:2|²)|m2|m²|m\?|\bmz\b)"

def _norm_num(n: str) -> str:
    if not n:
        return ""
    s = n.strip()
    s = s.replace(" ", "")
    s = s.replace(".", "")
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9.]", "", s)
    return s


def extract_total_area(text: str) -> str:
    t = re.sub(r"\s+", " ", (text or ""))
    if not t:
        return ""

    candidates: List[Tuple[int, float, str]] = []

    def add(score: int, num: str):
        nn = _norm_num(num)
        if not nn:
            return
        try:
            val = float(nn)
        except Exception:
            return
        candidates.append((score, val, nn))

    for m in re.finditer(rf"odnosno\s+ukupno\s*([0-9\.,]+)\s*{AREA_UNIT_RE}", t, flags=re.IGNORECASE):
        add(100, m.group(1))

    for m in re.finditer(rf"\bukupno\s*([0-9\.,]+)\s*{AREA_UNIT_RE}", t, flags=re.IGNORECASE):
        add(95, m.group(1))

    for m in re.finditer(rf"ukupn\w*\s+površin\w*\s*([0-9\.,]+)\s*{AREA_UNIT_RE}", t, flags=re.IGNORECASE):
        add(90, m.group(1))

    for m in re.finditer(rf"u\s+zakup.{0,80}?\bukupno\s*([0-9\.,]+)\s*{AREA_UNIT_RE}", t, flags=re.IGNORECASE):
        add(88, m.group(1))

    for m in re.finditer(rf"površin\w*\s*(?:od|:)?\s*([0-9\.,]+)\s*{AREA_UNIT_RE}", t, flags=re.IGNORECASE):
        add(70, m.group(1))

    for m in re.finditer(rf"u\s+površin\w*\s+od\s*([0-9\.,]+)\s*{AREA_UNIT_RE}", t, flags=re.IGNORECASE):
        add(65, m.group(1))

    if not candidates:
        return ""

    candidates.sort(key=lambda x: (x[0], x[1]), reverse=True)
    return candidates[0][2]


def extract_split_areas(text: str) -> Tuple[str, str, str]:
    t = re.sub(r"\s+", " ", (text or ""))
    if not t:
        return "", "", ""

    def grab(pattern: str) -> str:
        m = re.search(pattern, t, flags=re.IGNORECASE)
        if not m:
            return ""
        return _norm_num(m.group(1))

    skl = grab(rf"(?:skladiš\w*|skladište)\D{{0,50}}?([0-9\.,]+)\s*{AREA_UNIT_RE}")
    ured = grab(rf"(?:uredsk\w*|ured(?:i|a|e)?)\D{{0,50}}?([0-9\.,]+)\s*{AREA_UNIT_RE}")
    ost = grab(rf"(?:ostale?\s+prostor\w*|pomoćn\w*|sanitar\w*|garderob\w*)\D{{0,50}}?([0-9\.,]+)\s*{AREA_UNIT_RE}")

    return skl, ured, ost


def extract_notice_period(text: str) -> str:
    m = re.search(r"\botkazni\s+rok\s+od\s+([0-9]+)\s*dana\b", text, re.IGNORECASE)
    if m:
        return f"{m.group(1)} dana"
    return ""


CURRENCY_EUR_RE = r"(?:€|eur|eura|euro|eur-a)"
CURRENCY_HRK_RE = r"(?:kn|kuna|kune|hrk)"
M2_RE = r"(?:m\s*(?:2|²)|m2|m²|m\?)"

from decimal import Decimal, InvalidOperation

def _norm_money(num: str) -> str:
    if not num:
        return ""

    s = num.strip().replace(" ", "")
    s = s.replace("\u00a0", "")
    s = re.sub(r"[^0-9\.,]", "", s)
    if not s:
        return ""

    if "." in s and "," in s:
        s = s.replace(".", "")
        s = s.replace(",", ".")
    elif "," in s:
        s = s.replace(".", "")
        s = s.replace(",", ".")
    elif "." in s:
        left, right = s.rsplit(".", 1)
        if right.isdigit() and len(right) == 2:
            s = left.replace(".", "") + "." + right
        else:
            s = s.replace(".", "")

    try:
        val = Decimal(s)
    except InvalidOperation:
        return ""

    if val == val.to_integral_value():
        i = int(val)
        return f"{i:,}".replace(",", ".")

    val2 = val.quantize(Decimal("0.01"))
    whole = int(val2)
    dec = int((val2 - Decimal(whole)) * 100)

    whole_hr = f"{whole:,}".replace(",", ".")
    return f"{whole_hr},{dec:02d}"

def _first_money_after(pattern: str, text: str) -> str:
    m = re.search(pattern, text, flags=re.IGNORECASE)
    if not m:
        return ""
    return _norm_money(m.group(1))

def extract_unit_price(text: str) -> str:
    t = re.sub(r"\s+", " ", (text or ""))

    patterns = [
        rf"\bmjesečn\w*\s+zakupnin\w*.{{0,80}}?\biznosi\s*[:\-]?\s*([0-9\.,]+)\s*(?:{CURRENCY_EUR_RE}|{CURRENCY_HRK_RE})\s*(?:po\s*{M2_RE}|/{M2_RE})",
        rf"\biznosi\s+ukupno\s*([0-9\.,]+)\s*(?:{CURRENCY_EUR_RE}|{CURRENCY_HRK_RE})\s*(?:po\s*{M2_RE}|/{M2_RE})",
        rf"\b([0-9\.,]+)\s*(?:{CURRENCY_EUR_RE}|{CURRENCY_HRK_RE})\s*(?:po\s*{M2_RE}|/{M2_RE})\b",
    ]

    for p in patterns:
        v = _first_money_after(p, t)
        if v:
            return v

    return ""

def extract_monthly_rent(text: str) -> str:
    t = re.sub(r"\s+", " ", (text or ""))

    patterns = [
        rf"\bukupn\w*\s+mjesečn\w*\s+iznos\w*\s+od\s*([0-9\.,]+)\s*(?:{CURRENCY_EUR_RE}|{CURRENCY_HRK_RE})",
        rf"\bvisin\w*\s+mjesečn\w*\s+zakupnin\w*\s+iznosi\s*([0-9\.,]+)\s*(?:{CURRENCY_EUR_RE}|{CURRENCY_HRK_RE})",
        rf"\bmjesečn\w*\s+zakupnin\w*.{{0,140}}?\bšto\s+iznosi\s*(?:ukupno\s*)?([0-9\.,]+)\s*(?:{CURRENCY_EUR_RE}|{CURRENCY_HRK_RE})",
        rf"\bzakupnin\w*.{{0,60}}?\biznosi\s*([0-9\.,]+)\s*(?:{CURRENCY_EUR_RE}|{CURRENCY_HRK_RE}).{{0,60}}?\bmjesečn\w*\b",
    ]

    for p in patterns:
        v = _first_money_after(p, t)
        if v:
            return v

    return ""

def extract_amount_in_kunas(text: str) -> str:
    t = re.sub(r"\s+", " ", (text or ""))

    patterns = [
        rf"\bugovoren\w*.{{0,30}}?\b(?:u\s+kun\w*|{CURRENCY_HRK_RE})\b.{{0,30}}?([0-9\.,]+)\s*(?:{CURRENCY_HRK_RE})",
        rf"\bprotuvrijednost\s+u\s+kun\w*.{{0,40}}?([0-9\.,]+)\s*(?:{CURRENCY_HRK_RE})",
        rf"\bzakupnin\w*.{{0,60}}?([0-9\.,]+)\s*(?:{CURRENCY_HRK_RE})",
    ]

    for p in patterns:
        v = _first_money_after(p, t)
        if v:
            return v

    return ""

def extract_podzakup_flags(text: str) -> Tuple[str, str]:
    podzakup = ""
    suglasnost = ""
    if re.search(r"\bpodzakup\b", text, re.IGNORECASE):
        podzakup = "DA"
    if re.search(r"prethodna\s+pisana\s+suglasnost\s+zakupodavca", text, re.IGNORECASE):
        suglasnost = "DA"
    return podzakup, suglasnost

def extract_rezije_included(text: str) -> str:
    t = re.sub(r"\s+", " ", (text or ""))

    yes_patterns = [
        r"\b(u\s+trošak\s+zakupnine|u\s+zakupninu|u\s+cijenu\s+zakupnine)\s+uključeni\s+su\b.{0,80}?\brežijsk\w*\s+troškov\w*\b",
        r"\brežijsk\w*\s+troškov\w*\b.{0,80}?\b(uključeni\s+su|su\s+uključeni|uključuj\w*)\b.{0,40}?\b(zakupnin\w*|zakup)\b",
        r"\btroškov\w*\s+poslovanj\w*\b.{0,80}?\b(režijsk\w*\s+troškov\w*)\b.{0,80}?\b(uključeni\s+su|su\s+uključeni)\b",
    ]
    for p in yes_patterns:
        if re.search(p, t, flags=re.IGNORECASE):
            return "DA"

    no_patterns = [
        r"\b(po\s+potrošnj\w*|prema\s+potrošnj\w*)\b",
        r"\b(zakupnik\s+(snosi|plaća|podmiruje)|dužan\s+je\s+plaćati)\b.{0,120}?\b(režij\w*|troškov\w*\s+poslovanj\w*|komunal\w*|struj\w*|vod\w*|plin\w*|čišćenj\w*)\b",
        r"\bpored\s+zakupnine\b.{0,120}?\b(plaćat\w*|snosi|podmiruj\w*)\b",
        r"\b(režij\w*|komunal\w*|struj\w*|vod\w*|plin\w*)\b.{0,80}?\bnisu\s+uključeni\b",
        r"\bzasebno\s+se\s+plać\w*\b.{0,80}?\b(režij\w*|komunal\w*|troškov\w*)\b",
    ]
    for p in no_patterns:
        if re.search(p, t, flags=re.IGNORECASE):
            return "NE"

    return ""

def extract_zaduznica_amount(text: str) -> str:
    t = re.sub(r"\s+", " ", (text or ""))

    ZAD_RE = r"z\W*a\W*d\W*u\W*[žz]\W*n\W*i\W*c"

    if not re.search(ZAD_RE, t, flags=re.IGNORECASE):
        return ""

    BETWEEN_NUM_AND_CURR = r"(?:\s*\([^)]{0,40}\)\s*)?"

    patterns = [
        rf"\bna\s+(?:najviš\w*|največ\w*)\s+iznos\s+(?:od\s*)?([0-9\.,\s]+){BETWEEN_NUM_AND_CURR}\s*(?:{CURRENCY_EUR_RE}|{CURRENCY_HRK_RE})\b",
        rf"\bu\s+iznosu\s+(?:od\s*)?([0-9\.,\s]+){BETWEEN_NUM_AND_CURR}\s*(?:{CURRENCY_EUR_RE}|{CURRENCY_HRK_RE})\b",
        rf"{ZAD_RE}\w*.{{0,180}}?\biznos\s+(?:od\s*)?([0-9\.,\s]+){BETWEEN_NUM_AND_CURR}\s*(?:{CURRENCY_EUR_RE}|{CURRENCY_HRK_RE})\b",
        rf"{ZAD_RE}\w*.{{0,220}}?([0-9\.,\s]+){BETWEEN_NUM_AND_CURR}\s*(?:{CURRENCY_EUR_RE}|{CURRENCY_HRK_RE})\b",
    ]

    for p in patterns:
        m = re.search(p, t, flags=re.IGNORECASE)
        if m:
            return _norm_money(m.group(1))

    return ""

def extract_predmet_snippet(text: str) -> str:
    m = re.search(r"Predmet\s+zakupa(.{0,700})", text, re.IGNORECASE | re.DOTALL)
    if not m:
        return ""
    snippet = normalize_spaces(m.group(1))
    parts = re.split(r"(?<=[\.\!])\s+", snippet)
    return " ".join(parts[:2]).strip()


# =========================================================
# DATUMI (NOVA LOGIKA)
# =========================================================
DATE_RE = re.compile(r"\b([0-3]?\d)\s*[.\-\/]\s*([01]?\d)\s*[.\-\/]\s*((?:19|20)\d{2})\b")

def norm_date(d: str) -> str:
    m = DATE_RE.search(d)
    if not m:
        return ""
    dd = int(m.group(1))
    mm = int(m.group(2))
    yyyy = int(m.group(3))
    return f"{dd:02d}.{mm:02d}.{yyyy:04d}"

def extract_date_signed(text: str) -> str:
    t = re.sub(r"\s+", " ", (text or "")).strip()
    if not t:
        return ""

    chunks = [t[:HEAD_CHARS], t[-HEAD_CHARS:]]

    dpat = r"([0-3]?\d\s*[.\-\/]\s*[01]?\d\s*[.\-\/]\s*(?:19|20)\d{2})"

    patterns = [
        rf"\bU\s+[A-Za-zČĆŠĐŽčćšđž]+,\s*(?:dana\s*)?{dpat}\s*(?:godine)?\.?\b",
        rf"\bsklopljen\w*\s*{dpat}\b",
        rf"\bdana\s*{dpat}\s*(?:godine)?\.?\b",
    ]

    for c in chunks:
        for p in patterns:
            m = re.search(p, c, flags=re.IGNORECASE)
            if m:
                return norm_date(m.group(1))

    all_tail = list(DATE_RE.finditer(chunks[-1]))
    if all_tail:
        return norm_date(all_tail[-1].group(0))

    m2 = DATE_RE.search(chunks[0])
    return norm_date(m2.group(0)) if m2 else ""

def extract_start_end_dates(text: str, signed_date: str = "") -> Tuple[str, str]:
    t = re.sub(r"\s+", " ", (text or ""))

    m = re.search(
        r"(početk\w*\s+zakup\w*\s+od|počinje\s+sa|počinje\s+od|stupa\s+na\s+snagu\s+od)\s*"
        r"([0-3]?\d\s*[.\-\/]\s*[01]?\d\s*[.\-\/]\s*(?:19|20)\d{2})"
        r".{0,140}?"
        r"(završetk\w*\s+zakup\w*\s+dana|zakup\s+završava|završava|do)\s*"
        r"([0-3]?\d\s*[.\-\/]\s*[01]?\d\s*[.\-\/]\s*(?:19|20)\d{2})",
        t,
        flags=re.IGNORECASE
    )
    if m:
        return norm_date(m.group(2)), norm_date(m.group(4))

    m = re.search(
        r"(na\s+neodređen\w*).{0,80}?(počinje\s+sa|počinje\s+od)\s*"
        r"([0-3]?\d\s*[.\-\/]\s*[01]?\d\s*[.\-\/]\s*(?:19|20)\d{2})",
        t,
        flags=re.IGNORECASE
    )
    if m:
        return norm_date(m.group(3)), ""

    if re.search(r"aneks.*stupa\s+na\s+snagu.*danom\s+njegovog\s+potpisivanja", t, flags=re.IGNORECASE):
        return signed_date or "", ""

    m = re.search(r"stupa\s+na\s+snagu.{0,80}?" + DATE_RE.pattern, t, flags=re.IGNORECASE)
    if m:
        dm = DATE_RE.search(m.group(0))
        return norm_date(dm.group(0)) if dm else "", ""

    return "", ""

def extract_duration_months(text: str) -> str:
    t = re.sub(r"\s+", " ", (text or ""))

    if re.search(r"\bna\s+neodređen\w*\b", t, flags=re.IGNORECASE):
        return ""

    m = re.search(
        r"\b(rok|razdoblje)\s+od\s+(\d+)\s*(?:\([^)]*\))?\s*godin",
        t,
        flags=re.IGNORECASE
    )
    if m:
        try:
            return str(int(m.group(2)) * 12)
        except Exception:
            return ""

    m = re.search(
        r"\bna\s+(rok|razdoblje)\s+od\s+(\d+)\s*(?:\([^)]*\))?\s*godin",
        t,
        flags=re.IGNORECASE
    )
    if m:
        try:
            return str(int(m.group(2)) * 12)
        except Exception:
            return ""

    m = re.search(
        r"\btrajanj\w*.{0,120}?\bod\s+(\d+)\s*(?:\([^)]*\))?\s*godin",
        t,
        flags=re.IGNORECASE
    )
    if m:
        try:
            return str(int(m.group(1)) * 12)
        except Exception:
            return ""

    return ""


# =========================================================
# EXTRACT FIELDS (glavno)
# =========================================================
def extract_fields_from_text(text: str) -> Dict[str, Any]:
    out: Dict[str, Any] = {}

    z = extract_zakupodavac(text)
    if z["name"]:
        out["Naziv zakupodavca"] = z["name"]
    if z["oib"]:
        out["OIB zakupodavca"] = z["oib"]
    if z["address"]:
        out["Adresa sjedišta zakupodavca"] = z["address"]

    out["z.k.ul"] = extract_zk_ul(text)
    out["z.k.č.br."] = extract_zk_c_br(text)
    out["općinski sud"] = extract_opcinski_sud(text)

    out["katastarska općina"] = ""

    signed = extract_date_signed(text)
    start, end = extract_start_end_dates(text, signed_date=signed)

    out["Datum potpisivanja ugovora /aneksa"] = signed
    out["Datum stupanja ugovora na snagu"] = start
    out["Datum isteka ugovora"] = end
    out["Trajanje ugovora u mjesecima"] = extract_duration_months(text)

    out["otkaz"] = extract_notice_period(text)

    out["Kvadratura ukupnog prostora u zakupu u m2"] = extract_total_area(text)

    skl, ured, ost = extract_split_areas(text)
    out["Kvadratura skladišta"] = skl
    out["Kvadratura uredski prostori"] = ured
    out["Ostale prostorije"] = ost
    out["Broj parkirnih mjesta"] = extract_parking_first_item(text)

    out["Jedinična cijena/m2 (Bez PDV-a)"] = extract_unit_price(text)
    mj = extract_monthly_rent(text)

    out["AKTUALNA CIJENA ZAKUPA"] = mj
    out["Ukupna vrijednost ugovora (mjesečni iznos zakupnine)"] = mj
    out["Ukupna vrijednost ugovora ugovoreno u kunama"] = extract_amount_in_kunas(text)

    podz, sug = extract_podzakup_flags(text)
    if podz:
        out["Podzakup"] = podz
    if sug:
        out["Suglasnost za podzakup"] = sug

    rez = extract_rezije_included(text)
    if rez:
        out["Režijski troškovi uključeni?"] = rez

    zad = extract_zaduznica_amount(text)
    if zad:
        out["Zadužnice"] = zad

    out["datum za slanje maila za pregovaranje"] = ""

    return out


def build_row_dict(pdf_path: Path, text: str) -> Dict[str, Any]:
    filename = pdf_path.name
    loc = extract_location_from_filename(filename)

    row: Dict[str, Any] = {h: "" for h in EXCEL_HEADERS}

    row["Broj/naziv ugovora"] = detect_contract_title(text)
    row["Ugovor / aneks"] = detect_contract_type(text)
    row["odjel kod"] = loc or ""

    if loc and loc in LOCATION_MAP:
        m = LOCATION_MAP[loc]
        partner_name = m.get("PARTNER", "")

        row["Partner (podzakupnik)"] = partner_name
        row["Ime"] = m.get("Vlasnik", "")
        row["Adresa partnera"] = m.get("Adresa partnera", "")
        row["Poslovnica"] = m.get("Poslovnica", "")
        row["Adresa poslovnice"] = m.get("Adresa poslovnice", "")
        row["RVP"] = m.get("RVP", "")

        clean_partner = partner_name.replace(".", "").strip().upper()
        if clean_partner == "INTER CARS DOO":
            row["Korisnik ugovora"] = "UPRAVA (zakup)"
        else:
            row["Korisnik ugovora"] = "PARTNER (podzakup)"

    extracted = extract_fields_from_text(text)
    for k, v in extracted.items():
        if k in row and v is not None:
            row[k] = v

    return row


# =========================================================
# LIST PDF
# =========================================================
def list_pdfs(pdf_dir: Path) -> List[Path]:
    if not pdf_dir.exists():
        return []
    return sorted([p for p in pdf_dir.rglob("*.pdf") if p.is_file()])

# =========================================================
# NOVO: Upload helper (drag & drop ZA DATOTEKE)
# =========================================================
def get_upload_tmp_dir() -> Path:
    if "upload_tmp_dir" not in st.session_state:
        st.session_state["upload_tmp_dir"] = Path(tempfile.mkdtemp(prefix="zakupi_upload_"))
    return Path(st.session_state["upload_tmp_dir"])

def save_uploaded_pdfs(uploaded_files) -> List[Path]:
    tmp_dir = get_upload_tmp_dir()
    saved: List[Path] = []
    for uf in uploaded_files or []:
        name = Path(uf.name).name
        out_path = tmp_dir / name
        with open(out_path, "wb") as f:
            f.write(uf.getbuffer())
        saved.append(out_path)
    return saved

# =========================================================
# STREAMLIT UI
# =========================================================
st.set_page_config(page_title="Zakupi: OCR PDF → Zakup.xlsx", layout="wide")
st.title("📄 Obrada ugovora o zakupu (OCR)")

with st.sidebar:
    st.subheader("Izvor PDF-ova")
    source_mode = st.radio(
        "Odabir izvora",
        ["Folder (server path)", "Upload datoteka (drag & drop)"],
        index=0
    )

    st.subheader("Postavke outputa")
    pdf_dir = Path(st.text_input("Mapa s PDF-ovima", DEFAULT_PDF_DIR))
    excel_path = Path(st.text_input("Excel output", DEFAULT_EXCEL_PATH))
    sheet_name = st.text_input("Sheet name", DEFAULT_SHEET)

    st.divider()
    st.caption("Svi dokumenti su skenirani: OCR se uvijek koristi.")

pdfs: List[Path] = []
source_label = ""

if source_mode == "Folder (server path)":
    pdfs = list_pdfs(pdf_dir)
    source_label = f"Folder: `{pdf_dir}`"
else:
    with st.sidebar:
        st.caption("⬇️ Uploadaj PDF-ove (drag & drop datoteka u okvir ispod)")
        uploaded = st.file_uploader(
            "PDF upload",
            type=["pdf"],
            accept_multiple_files=True
        )
    if uploaded:
        pdfs = save_uploaded_pdfs(uploaded)
        source_label = f"Upload: {len(pdfs)} datoteka"
    else:
        pdfs = []
        source_label = "Upload: (nema datoteka)"

st.write(f"**Izvor:** {source_label}")
st.write(f"Pronađeno PDF-ova: **{len(pdfs)}**")

if not pdfs:
    st.stop()

names = [p.name for p in pdfs]
selected = st.multiselect(
    "Odaberi PDF-ove za obradu",
    options=names,
    default=names[:10] if len(names) > 10 else names
)

colA, colB = st.columns([1, 1])
with colA:
    process_btn = st.button("Obradi odabrane", type="primary")
with colB:
    preview_btn = st.button("Preview prvog odabranog")

if preview_btn and selected:
    p = next(x for x in pdfs if x.name == selected[0])
    t = read_pdf_scanned(p)

    st.subheader(f"Preview: {p.name}")
    st.text_area("Extracted tekst (OCR)", t, height=350)

    row = build_row_dict(p, t)  # build_row_dict mora ostati iz tvog originalnog koda
    st.subheader("Extracted polja (što ide u Excel)")
    st.json(row)

if process_btn:
    if not selected:
        st.warning("Nisi odabrao nijedan PDF.")
        st.stop()

    ensure_excel(excel_path, sheet_name)

    processed_rows = []
    prog = st.progress(0)
    sel_paths = [p for p in pdfs if p.name in selected]
    total = len(sel_paths)

    for i, p in enumerate(sel_paths, start=1):
        text = read_pdf_scanned(p)

        row = build_row_dict(p, text)
        pdf_web = onedrive_web_url_for_path(p)

        new_row_idx = append_row(
            excel_path,
            sheet_name,
            row,
            hyperlinks={"Broj/naziv ugovora": pdf_web}
        )

        processed_rows.append({
            "PDF": p.name,
            "ExcelRow": new_row_idx,
            "odjel kod": row.get("odjel kod", ""),
            "Poslovnica": row.get("Poslovnica", ""),
            "Partner": row.get("Partner (podzakupnik)", ""),
            "Zakupodavac": row.get("Naziv zakupodavca", ""),
            "Datum potpisivanja": row.get("Datum potpisivanja ugovora /aneksa", ""),
            "Stupa na snagu": row.get("Datum stupanja ugovora na snagu", ""),
            "Istek": row.get("Datum isteka ugovora", ""),
            "Trajanje(mj)": row.get("Trajanje ugovora u mjesecima", ""),
        })

        prog.progress(i / total)

    st.success(f"Obrađeno: {len(processed_rows)} PDF-ova. Upisano u: {excel_path}")
    st.subheader("Sažetak")
    st.dataframe(processed_rows, use_container_width=True)

    if not OCR_AVAILABLE:
        st.warning("OCR nije dostupan (nedostaje pytesseract/pdf2image). Instaliraj potrebne pakete.")
